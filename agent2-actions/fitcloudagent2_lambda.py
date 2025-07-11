import json
import os
import boto3
import requests
import logging
import re
from typing import Dict, Any
import openpyxl
from openpyxl.chart import BarChart, Reference
import io

logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Agent1 람다 이름 (슈퍼바이저가 처리하므로 선택사항)
AGENT1_LAMBDA_NAME = os.environ.get("AGENT1_LAMBDA_NAME", "fitcloud_action_part1-wpfe6")

# Agent1 ID와 Alias (Agent2에서 직접 호출할 때 사용)
AGENT1_ID = os.environ.get("AGENT1_ID", "NBLVKZOU76")
AGENT1_ALIAS = os.environ.get("AGENT1_ALIAS", "PSADGJ398L")

# 슬랙 토큰/채널ID를 환경변수에서 가져오기 (보안상 하드코딩 금지)
SLACK_BOT_TOKEN = os.environ.get('SLACK_BOT_TOKEN')
SLACK_CHANNEL = os.environ.get('SLACK_CHANNEL')

# Bedrock 클라이언트 초기화
bedrock_client = boto3.client('bedrock-runtime')

# 환경변수 검증
if not SLACK_BOT_TOKEN:
    raise ValueError("SLACK_BOT_TOKEN 환경변수가 설정되지 않았습니다.")
if not SLACK_CHANNEL:
    raise ValueError("SLACK_CHANNEL 환경변수가 설정되지 않았습니다.")

def parse_agent1_response_with_llm(input_text: str) -> list:
    """
    LLM을 사용해서 Agent1의 응답을 구조화된 데이터로 변환합니다.
    """
    try:
        logger.info(f"[Agent2] LLM 파싱 시작 - 입력 길이: {len(input_text)}")
        
        # 이스케이프 문자 처리 개선
        try:
            # 먼저 일반적인 이스케이프 문자 처리
            decoded_text = input_text.encode('utf-8').decode('unicode_escape')
            logger.info(f"[Agent2] 이스케이프 문자 처리 후 (처음 300자): {decoded_text[:300]}")
            input_text = decoded_text
        except Exception as e:
            logger.warning(f"[Agent2] 이스케이프 문자 처리 실패: {e}")
            # 실패 시 원본 텍스트 사용
            logger.info(f"[Agent2] 원본 텍스트 사용 (처음 300자): {input_text[:300]}")
        
        # LLM에게 파싱 요청 (더 강력한 프롬프트)
        prompt = f"""
다음은 AWS 비용/사용량 조회 결과입니다. 이 텍스트를 분석해서 엑셀 파일에 적합한 구조화된 데이터로 변환해주세요.

**중요**: 모든 서비스 항목을 누락 없이 추출해야 합니다. 텍스트에 언급된 모든 서비스와 금액을 포함하세요.

요구사항:
1. 텍스트에 언급된 **모든 서비스**를 추출 (누락 금지)
2. 각 서비스의 이름, 금액, 비율을 정확히 추출
3. JSON 배열 형태로 반환
4. 각 항목은 serviceName, usageFeeUSD, percentage, billingPeriod 필드를 포함
5. 월 정보가 있으면 billingPeriod에 YYYYMM 형식으로 포함
6. "기타 서비스"나 "기타" 항목도 별도로 포함
7. 총 38개 항목이 있다면 38개 모두 추출

**파싱 규칙**:
- "**서비스명**: $금액 (비율%)" 패턴 추출
- "기타 서비스: $금액 (비율%)" 패턴도 추출
- 모든 숫자와 비율을 정확히 포함
- 서비스명에 특수문자(*, -, 등)가 있어도 그대로 포함
- 로그 정보나 디버그 정보는 무시하고 실제 데이터만 추출

**데이터 추출 우선순위**:
1. [RESPONSE][message] 섹션의 데이터
2. 마크다운 형식의 서비스별 데이터
3. 기타 서비스 정보

**예시 패턴**:
- **AmazonRDS**: $10,041.24 (46.5%)
- **Saltware Care Pack (FR)**: $4,726.33 (21.9%)
- **AmazonEC2**: $2,086.94 (9.7%)
- **기타 서비스**: $1,584.74 (7.3%)

입력 텍스트:
{input_text}

응답은 반드시 JSON 배열 형태로만 반환하세요. 다른 설명이나 텍스트는 포함하지 마세요.
모든 서비스를 누락 없이 포함해야 합니다.
"""

        logger.info(f"[Agent2] Bedrock LLM 호출 시작")
        
        # Bedrock LLM 호출 (타임아웃 설정 단축)
        import botocore
        
        config = botocore.config.Config(
            read_timeout=120,  # 2분으로 단축
            connect_timeout=30  # 30초로 단축
        )
        
        bedrock_client_with_timeout = boto3.client('bedrock-runtime', config=config)
        
        response = bedrock_client_with_timeout.invoke_model(
            modelId='anthropic.claude-3-5-sonnet-20240620-v1:0',
            body=json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 8000,  # 토큰 제한 증가
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ]
            })
        )
        
        logger.info(f"[Agent2] Bedrock LLM 응답 수신")
        
        response_body = json.loads(response['body'].read())
        llm_response = response_body['content'][0]['text']
        logger.info(f"[Agent2] LLM 응답 (처음 300자): {llm_response[:300]}")
        logger.info(f"[Agent2] LLM 응답 전체 길이: {len(llm_response)}")
        
        # JSON 파싱
        try:
            # JSON 코드블록이 있으면 추출
            json_match = re.search(r'```json\s*(\[.*?\])\s*```', llm_response, re.DOTALL)
            if json_match:
                json_str = json_match.group(1)
                logger.info(f"[Agent2] JSON 코드블록 추출 성공 (길이: {len(json_str)})")
                parsed_data = json.loads(json_str)
            else:
                # 직접 JSON 파싱 시도
                logger.info(f"[Agent2] 직접 JSON 파싱 시도")
                parsed_data = json.loads(llm_response)
            
            logger.info(f"[Agent2] LLM 파싱 성공: {len(parsed_data)}개 항목")
            
            # 파싱된 데이터 로그 (모든 항목)
            for i, item in enumerate(parsed_data):
                logger.info(f"[Agent2] 파싱된 항목 {i+1}: {item}")
            
            # 파싱 결과 요약
            total_amount = sum(item.get('usageFeeUSD', 0) for item in parsed_data)
            logger.info(f"[Agent2] 파싱된 총 금액: ${total_amount:,.2f}")
            logger.info(f"[Agent2] 파싱된 항목 수: {len(parsed_data)}개")
            
            return parsed_data
            
        except json.JSONDecodeError as e:
            logger.error(f"[Agent2] LLM 응답 JSON 파싱 실패: {e}")
            logger.error(f"[Agent2] LLM 응답 전체: {llm_response}")
            return []
            
    except Exception as e:
        logger.error(f"[Agent2] LLM 파싱 중 오류: {e}")
        import traceback
        logger.error(f"[Agent2] LLM 파싱 오류 상세: {traceback.format_exc()}")
        
        # LLM 실패 시 기본 데이터 구조로 변환 시도
        try:
            logger.info(f"[Agent2] LLM 실패, 기본 파싱 시도")
            # Agent1 응답에서 숫자와 서비스명 추출
            
            # 총 금액 추출
            total_match = re.search(r'총 온디맨드 사용금액: \$([0-9,]+\.?\d*)', input_text)
            total_amount = float(total_match.group(1).replace(',', '')) if total_match else 0
            
            # 서비스별 데이터 추출
            services = []
            service_pattern = r'(\d+)\. \*?([^*]+)\*?: 약 \$([0-9,]+) \(([0-9.]+)%\)'
            matches = re.findall(service_pattern, input_text)
            
            for rank, service_name, amount, percentage in matches:
                services.append({
                    "serviceName": service_name.strip(),
                    "usageFeeUSD": float(amount.replace(',', '')),
                    "percentage": float(percentage),
                    "billingPeriod": datetime.now().strftime("%Y%m")  # 현재 월을 기본값으로
                })
            
            if services:
                logger.info(f"[Agent2] 기본 파싱 성공: {len(services)}개 서비스")
                return services
            
        except Exception as fallback_e:
            logger.error(f"[Agent2] 기본 파싱도 실패: {fallback_e}")
        
        return []

def generate_excel_report(data):
    """
    데이터를 받아서 엑셀 보고서를 생성하고 슬랙에 업로드하는 함수
    """
    # app.py에서 개선된 데이터 검증 로직 적용
    if not data or not isinstance(data, list) or len(data) == 0:
        raise ValueError("유효하지 않은 데이터입니다. 리스트 형태의 데이터가 필요합니다.")

    records = data
    first = records[0]
    
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 데이터 구조 자동 판별 (app.py와 동일한 로직)
    excel_title = "AWS 리포트"
    ws_title = "리포트"
    headers = []
    rows = []
    chart = None
    chart_x_title = ''
    chart_y_title = ''
    chart_title = ''

    # inputText에서 추출한 가상 데이터 구조 처리 (새로 추가)
    if 'percentage' in first and 'billingPeriod' in first:
        ws_title = "서비스별 요금 리포트"
        headers = ['순위', '서비스명', '요금($)', '비율(%)']
        rows = []
        for i, item in enumerate(records, 1):
            rows.append([
                i,
                item.get('serviceName', ''),
                item.get('usageFeeUSD', 0),
                item.get('percentage', 0)
            ])
        chart_x_title = '서비스명'
        chart_y_title = '요금(USD)'
        chart_title = '서비스별 요금'
    # 월별 요금
    elif 'billingPeriod' in first:
        ws_title = "월별 요금 리포트"
        headers = ['월', '요금($)']
        months = [item['billingPeriod'] for item in records]
        costs = [float(item.get('usageFee', item.get('usageFeeUSD', 0))) for item in records]
        rows = list(zip(months, costs))
        chart_x_title = '월'
        chart_y_title = '요금(USD)'
        chart_title = '월별 요금'
    # 일별 요금
    elif 'date' in first or 'dailyDate' in first:
        ws_title = "일별 요금 리포트"
        headers = ['일', '요금($)']
        days = [item.get('date', item.get('dailyDate')) for item in records]
        costs = [float(item.get('usageFee', item.get('usageFeeUSD', 0))) for item in records]
        rows = list(zip(days, costs))
        chart_x_title = '일'
        chart_y_title = '요금(USD)'
        chart_title = '일별 요금'
    # 계정별 요금
    elif 'accountId' in first:
        ws_title = "계정별 요금 리포트"
        headers = ['계정ID', '요금($)']
        accounts = [item['accountId'] for item in records]
        costs = [float(item.get('usageFee', item.get('usageFeeUSD', 0))) for item in records]
        rows = list(zip(accounts, costs))
        chart_x_title = '계정ID'
        chart_y_title = '요금(USD)'
        chart_title = '계정별 요금'
    # 태그별 요금 등 기타 케이스(확장 가능) - app.py와 동일한 주석
    elif 'tagsJson' in first:
        ws_title = "태그별 요금 리포트"
        headers = ['태그', '요금($)']
        tags = []
        costs = []
        for item in records:
            tag_str = ', '.join([f'{k}:{v}' for k, v in item['tagsJson'].items()]) if isinstance(item['tagsJson'], dict) else str(item['tagsJson'])
            tags.append(tag_str)
            costs.append(float(item.get('usageFee', item.get('usageFeeUSD', 0))))
        rows = list(zip(tags, costs))
        chart_x_title = '태그'
        chart_y_title = '요금(USD)'
        chart_title = '태그별 요금'
    else:
        # 모든 필드를 헤더로, 각 row를 값으로
        headers = list(first.keys())
        rows = [[item.get(h, '') for h in headers] for item in records]
        ws_title = "일반 리포트"
        chart = None  # 차트 미생성

    ws.title = ws_title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    
    # 금액 컬럼에 통화 형식 적용
    if 'percentage' in first and 'billingPeriod' in first:
        # 서비스별 요금 리포트인 경우 3번째 컬럼(요금)에 통화 형식 적용
        for row_num in range(2, len(rows) + 2):  # 헤더 다음 행부터
            cell = ws.cell(row=row_num, column=3)  # 3번째 컬럼 (요금)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
    elif 'billingPeriod' in first:
        # 월별 요금 리포트인 경우 2번째 컬럼(요금)에 통화 형식 적용
        for row_num in range(2, len(rows) + 2):
            cell = ws.cell(row=row_num, column=2)  # 2번째 컬럼 (요금)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
    elif 'date' in first or 'dailyDate' in first:
        # 일별 요금 리포트인 경우 2번째 컬럼(요금)에 통화 형식 적용
        for row_num in range(2, len(rows) + 2):
            cell = ws.cell(row=row_num, column=2)  # 2번째 컬럼 (요금)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
    elif 'accountId' in first:
        # 계정별 요금 리포트인 경우 2번째 컬럼(요금)에 통화 형식 적용
        for row_num in range(2, len(rows) + 2):
            cell = ws.cell(row=row_num, column=2)  # 2번째 컬럼 (요금)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'
    elif 'tagsJson' in first:
        # 태그별 요금 리포트인 경우 2번째 컬럼(요금)에 통화 형식 적용
        for row_num in range(2, len(rows) + 2):
            cell = ws.cell(row=row_num, column=2)  # 2번째 컬럼 (요금)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'

    # 차트 추가 (가능한 경우만) - app.py와 동일한 로직
    if not chart and len(rows) > 0 and len(headers) >= 2:
        # LLM에서 추출한 데이터의 경우 3번째 컬럼(요금)을 차트 데이터로 사용
        if 'percentage' in first and 'billingPeriod' in first:
            chart = BarChart()
            chart.title = chart_title
            chart.x_axis.title = chart_x_title
            chart.y_axis.title = chart_y_title
            data_ref = Reference(ws, min_col=3, min_row=1, max_row=len(rows)+1)  # 3번째 컬럼 (요금)
            cats_ref = Reference(ws, min_col=2, min_row=2, max_row=len(rows)+1)  # 2번째 컬럼 (서비스명)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "F2")
        elif len(headers) == 2 and all(isinstance(r[1], (int, float)) for r in rows):
            chart = BarChart()
            chart.title = chart_title
            chart.x_axis.title = chart_x_title
            chart.y_axis.title = chart_y_title
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(rows)+1)
            cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(rows)+1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            ws.add_chart(chart, "E2")

    # 파일 메모리 저장
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    file_name = 'report.xlsx'
    mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    file_stream_value = file_stream.getvalue()
    file_size = len(file_stream_value)

    # 슬랙 파일 업로드 (app.py와 동일한 개선된 로직)
    try:
        headers_get_url = {
            'Authorization': f'Bearer {SLACK_BOT_TOKEN}'
        }
        files_data = {
            'filename': (None, file_name),
            'length': (None, str(file_size)),
            'filetype': (None, 'xlsx')
        }
        
        # 1. 업로드 URL 가져오기
        get_upload_url_response = requests.post(
            'https://slack.com/api/files.getUploadURLExternal',
            headers=headers_get_url,
            files=files_data,
            timeout=30  # 30초 타임아웃 추가
        )
        get_upload_url_result = get_upload_url_response.json()
        if not get_upload_url_result.get('ok'):
            error_msg = get_upload_url_result.get('error')
            raise Exception(f'파일 업로드 URL을 가져오는 데 실패했습니다: {error_msg}')
            
        upload_url = get_upload_url_result['upload_url']
        file_id = get_upload_url_result['file_id']
        
        # 2. 파일 콘텐츠 업로드
        file_stream.seek(0)
        files = {
            'file': (file_name, file_stream, mime_type)
        }
        upload_file_response = requests.post(
            upload_url,
            files=files,
            timeout=60  # 60초 타임아웃 추가
        )
        if not upload_file_response.ok:
            raise Exception(f'파일 콘텐츠 업로드에 실패했습니다: {upload_file_response.text}')
            
        # 3. 업로드 완료
        headers_complete_upload = {
            'Authorization': f'Bearer {SLACK_BOT_TOKEN}',
            'Content-Type': 'application/json'
        }
        payload_complete_upload = {
            'files': [{'id': file_id, 'title': file_name}],
            'channel_id': SLACK_CHANNEL,
            'initial_comment': f'📊 {ws_title}가 생성되었습니다.'
        }
        complete_upload_response = requests.post(
            'https://slack.com/api/files.completeUploadExternal',
            headers=headers_complete_upload,
            json=payload_complete_upload,
            timeout=30  # 30초 타임아웃 추가
        )
        complete_upload_result = complete_upload_response.json()
        if not complete_upload_result.get('ok'):
            error_msg = complete_upload_result.get('error')
            if error_msg == 'not_in_channel':
                raise Exception('봇이 채널에 추가되지 않았습니다. 슬랙 채널에 봇을 추가해주세요.')
            elif error_msg == 'channel_not_found':
                raise Exception('채널을 찾을 수 없습니다. 채널 ID를 확인해주세요.')
            else:
                raise Exception(f'파일 업로드를 완료하는 데 실패했습니다: {error_msg}')
                
        permalink = None
        if complete_upload_result.get('files') and len(complete_upload_result['files']) > 0:
            permalink = complete_upload_result['files'][0].get('permalink')
            
        return {
            'success': True,
            'message': '파일 업로드 및 채널 공유 성공',
            'file_id': file_id,
            'permalink': permalink,
            'report_title': ws_title
        }
        
    except requests.exceptions.RequestException as e:
        raise Exception(f'네트워크 요청 오류: {e}')
    except Exception as e:
        raise Exception(f'예상치 못한 오류 발생: {e}')

def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Agent2(리포트/엑셀/슬랙 업로드) 람다
    1. 입력 파라미터(기간, 계정, 태그 등) 파싱
    2. Agent1 람다 호출하여 표 데이터 조회
    3. 엑셀 보고서 생성 및 슬랙 업로드
    4. 결과 반환
    """
    try:
        logger.info(f"[Agent2] Agent2 람다 시작")
        
        # 1. 파라미터 추출
        params = None
        if isinstance(event.get("parameters"), dict):
            params = event["parameters"]
        if not params:
            try:
                props = event["requestBody"]["content"]["application/json"]["properties"]
                for prop in props:
                    if prop.get("name") == "user_input":
                        params = {"user_input": prop.get("value")}
                        break
            except Exception as e:
                logger.error(f"user_input 추출 실패: {e}")
        if not params:
            params = event.get("user_input") or event.get("inputText") or event
        if isinstance(params, str):
            try:
                params = json.loads(params)
            except Exception:
                params = {"user_input": params}
        logger.info(f"[Agent2] 입력 파라미터: {params}")

        # 2. Agent1 데이터 추출
        agent1_result = None
        
        # 2-1. inputText에서 Agent1 데이터 추출 (직접 호출된 경우)
        if 'inputText' in event:
            input_text = event['inputText']
            logger.info(f"[Agent2] inputText에서 Agent1 데이터 추출 시도 (길이: {len(input_text)})")
            logger.info(f"[Agent2] inputText 내용 (처음 300자): {input_text[:300]}")
            
            # LLM을 사용해서 Agent1 응답 파싱
            agent1_result = parse_agent1_response_with_llm(input_text)
            if agent1_result:
                logger.info(f"[Agent2] inputText에서 LLM 파싱 성공: {len(agent1_result)}개 항목")
                for i, item in enumerate(agent1_result[:3]):  # 처음 3개만 로그
                    logger.info(f"[Agent2] 항목 {i+1}: {item.get('serviceName', 'N/A')} - ${item.get('usageFeeUSD', 0)} ({item.get('percentage', 0)}%)")
        
        # 2-2. conversationHistory에서 Agent1 데이터 추출 (백업)
        if not agent1_result and 'conversationHistory' in event:
            logger.info(f"[Agent2] conversationHistory에서 Agent1 데이터 추출 시도")
            ch = event['conversationHistory']
            if isinstance(ch, dict) and 'messages' in ch:
                for msg in reversed(ch['messages']):
                    if msg.get('role') == 'assistant' and msg.get('content'):
                        content = msg.get('content', '')
                        if isinstance(content, list) and len(content) > 0:
                            content_text = content[0]
                        else:
                            content_text = str(content)
                        
                        logger.info(f"[Agent2] conversationHistory에서 assistant 메시지 발견 (길이: {len(content_text)})")
                        agent1_result = parse_agent1_response_with_llm(content_text)
                        if agent1_result:
                            logger.info(f"[Agent2] conversationHistory에서 LLM 파싱 성공: {len(agent1_result)}개 항목")
                            break
        
        # 3. 최종 검증 - Agent1 데이터가 없으면 Agent1을 직접 호출
        if not agent1_result:
            logger.warning('[Agent2] Agent1의 데이터를 찾을 수 없습니다. Agent1을 직접 호출합니다.')
            
            try:
                # Agent1을 직접 호출하여 데이터 조회
                logger.info('[Agent2] Agent1 직접 호출 시작')
                
                # Bedrock Agent Runtime 클라이언트
                client = boto3.client("bedrock-agent-runtime")
                
                # Agent1 ID와 Alias (환경변수에서 가져오기)
                agent1_id = AGENT1_ID
                agent1_alias = AGENT1_ALIAS
                session_id = event.get("sessionId", "agent2-fallback-session")
                
                # 사용자 입력에서 날짜/계정 정보 추출
                user_input = event.get("inputText", "")
                if not user_input:
                    # conversationHistory에서 마지막 사용자 요청 추출
                    if 'conversationHistory' in event:
                        ch = event['conversationHistory']
                        if isinstance(ch, dict) and 'messages' in ch:
                            for msg in reversed(ch['messages']):
                                if msg.get('role') == 'user':
                                    user_input = msg.get('content', '')
                                    if isinstance(user_input, list) and len(user_input) > 0:
                                        user_input = user_input[0]
                                    break
                
                if not user_input:
                    # 현재 날짜를 기반으로 기본값 설정
                    from datetime import datetime
                    current_date = datetime.now()
                    user_input = f"{current_date.year}년 {current_date.month}월 서비스별 사용량 조회"
                
                logger.info(f'[Agent2] Agent1 호출 파라미터: sessionId={session_id}, inputText={user_input}')
                
                # Agent1 호출
                agent1_response = client.invoke_agent(
                    agentId=agent1_id,
                    agentAliasId=agent1_alias,
                    sessionId=session_id,
                    inputText=user_input
                )
                
                # Agent1 응답 처리
                raw_agent1_response = ""
                for event_chunk in agent1_response:
                    if 'chunk' in event_chunk and 'bytes' in event_chunk['chunk']:
                        raw_agent1_response += event_chunk['chunk']['bytes'].decode('utf-8')
                
                logger.info(f'[Agent2] Agent1 응답 받음 (길이: {len(raw_agent1_response)})')
                
                # Agent1 응답을 Agent2에서 파싱
                agent1_result = parse_agent1_response_with_llm(raw_agent1_response)
                
                if agent1_result:
                    logger.info(f'[Agent2] Agent1 직접 호출로 데이터 획득 성공: {len(agent1_result)}개 항목')
                else:
                    logger.error('[Agent2] Agent1 직접 호출 후에도 데이터 파싱 실패')
                    return {
                        'response': {
                            'body': {
                                'content': [
                                    {
                                        'type': 'text',
                                        'text': '❌ Agent1에서 데이터를 조회할 수 없습니다. 잠시 후 다시 시도해주세요.'
                                    }
                                ]
                            }
                        }
                    }
                    
            except Exception as e:
                logger.error(f'[Agent2] Agent1 직접 호출 실패: {e}')
                return {
                    'response': {
                        'body': {
                            'content': [
                                {
                                    'type': 'text',
                                    'text': f'❌ Agent1 호출 중 오류가 발생했습니다: {str(e)}'
                                }
                            ]
                        }
                    }
                }
        
        logger.info(f"[Agent2] Agent1 데이터 추출 완료 - 타입: {type(agent1_result)}, 길이: {len(agent1_result) if isinstance(agent1_result, list) else 'N/A'}")

        # 4. 데이터 검증
        if not agent1_result:
            logger.error(f"[Agent2] Agent1 데이터가 없습니다")
            raise ValueError("Agent1의 데이터가 없습니다. 먼저 비용/사용량을 조회해주세요.")
        
        if not isinstance(agent1_result, list):
            logger.warning(f"[Agent2] Agent1 응답이 리스트가 아님: {type(agent1_result)}")
            agent1_result = []
        
        if len(agent1_result) == 0:
            logger.warning(f"[Agent2] Agent1 데이터가 비어있음")
            return {
                'response': {
                    'body': {
                        'content': [
                            {
                                'type': 'text',
                                'text': '📊 Agent1에서 조회된 데이터가 없습니다. 다른 조건으로 조회해보세요.'
                            }
                        ]
                    }
                }
            }

        logger.info(f"[Agent2] 데이터 검증 완료, 레코드 수: {len(agent1_result)}")

        # 5. 엑셀 보고서 생성 및 슬랙 업로드
        logger.info(f"[Agent2] 엑셀 보고서 생성 시작")
        try:
            upload_result = generate_excel_report(agent1_result)
            logger.info(f"[Agent2] 엑셀 보고서 생성 완료")
        except Exception as e:
            logger.error(f"[Agent2] 엑셀 보고서 생성 실패: {e}")
            import traceback
            logger.error(f"[Agent2] 엑셀 생성 실패 상세: {traceback.format_exc()}")
            raise

        # 6. 결과 반환 (Bedrock Agent Action Group 응답 형식)
        completion_msg = (
            f"📊 **{upload_result.get('report_title', '리포트')} 생성 완료!**\n"
            f"✅ 엑셀 파일이 슬랙 채널에 업로드되었습니다.\n"
            f"🔗 파일 링크: {upload_result.get('permalink', '링크 없음')}\n"
            f"📁 파일 ID: {upload_result.get('file_id', 'N/A')}\n"
            f"📋 데이터 소스: {'세션 속성' if 'sessionAttributes' in event and isinstance(event['sessionAttributes'], dict) and 'agent1_result' in event['sessionAttributes'] else 'Agent1 호출'}"
        )
        
        logger.info(f"[Agent2] Bedrock Agent Action Group 응답 반환")
        
        # Bedrock Agent Action Group이 기대하는 응답 형식
        return {
            'response': {
                'body': {
                    'content': [
                        {
                            'type': 'text',
                            'text': completion_msg
                        }
                    ]
                }
            }
        }

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        logger.error(f"[Agent2] 처리 중 오류: {e}\n{tb}", exc_info=True)
        
        # Bedrock Agent가 기대하는 응답 형식으로 오류 반환
        error_message = f"❌ [Agent2] 처리 중 오류가 발생했습니다: {str(e)}"
        
        return {
            'response': {
                'body': {
                    'content': [
                        {
                            'type': 'text',
                            'text': error_message
                        }
                    ]
                }
            }
        } 